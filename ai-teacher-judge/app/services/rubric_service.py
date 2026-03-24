from __future__ import annotations

import io
import json
from time import perf_counter
from typing import Any

import httpx
from fastapi import HTTPException

from app.core.config import settings
from app.schemas.rubric import ChatMessage, RubricAnalysis, RubricItem


# ──────────────────────────────────────────────────────────────
# Shared helpers
# ──────────────────────────────────────────────────────────────

def _strip_think_tags(text: str) -> str:
    """Keep only content after </think>; return text as-is if tag absent."""
    marker = "</think>"
    idx = text.find(marker)
    return text[idx + len(marker):].strip() if idx != -1 else text.strip()


def _apply_thinking_control(payload: dict[str, Any]) -> dict[str, Any]:
    payload["chat_template_kwargs"] = {
        **dict(payload.get("chat_template_kwargs") or {}),
        "enable_thinking": settings.vllm_enable_thinking,
    }
    return payload


def _vllm_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {settings.vllm_api_key}",
        "Content-Type": "application/json",
    }


async def _call_vllm(payload: dict[str, Any], timeout: float = 60.0) -> tuple[str, dict]:
    """Call vLLM chat/completions and return (content, usage_metrics)."""
    url = f"{settings.vllm_base_url}/chat/completions"
    started = perf_counter()
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(url, json=payload, headers=_vllm_headers())
            resp.raise_for_status()
            data = resp.json()

        elapsed = max(perf_counter() - started, 0.0)
        usage = data.get("usage") or {}
        prompt_tokens = int(usage.get("prompt_tokens") or 0)
        completion_tokens = int(usage.get("completion_tokens") or 0)
        total_tokens = int(usage.get("total_tokens") or (prompt_tokens + completion_tokens))
        tps = (completion_tokens / elapsed) if elapsed > 0 else 0.0

        content = data["choices"][0]["message"]["content"] or ""
        content = _strip_think_tags(content)
        metrics = {
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "total_tokens": total_tokens,
            "elapsed_seconds": round(elapsed, 3),
            "tokens_per_second": round(tps, 2),
        }
        return content, metrics
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"AI 呼叫失敗：{exc}") from exc


# ──────────────────────────────────────────────────────────────
# 1. Analyze rubric document
# ──────────────────────────────────────────────────────────────

_ANALYZE_SYSTEM_PROMPT = """
# 角色
你是一位專業的教學評分助理，服務對象是校園雲端平台的授課老師。

# 平台背景（內部知識）
本系統為校園雲端平台，學生作業在以下環境中執行：
- Proxmox 虛擬機器（VM）或 LXC 容器
- 本地部署，不使用公有雲（AWS/GCP/Azure）

## 系統可以自動偵測的資訊（透過 Proxmox Agent / API）
- 特定 TCP Port 是否正在監聽（e.g., Port 80/443/3306/5432）
- Linux 服務狀態（`systemctl status nginx` 等）
- 進程是否存在（`ps aux`）
- 磁碟/CPU/記憶體使用率
- 特定檔案是否存在（透過 Agent Exec）
- HTTP endpoint 回傳狀態碼

## 系統無法自動偵測的資訊（需人工判斷）
- 程式碼品質、架構設計、MVC 分層等
- 資料庫內容正確性（需帳密才能查詢）
- 圖形化介面截圖（需 VNC/螢幕截圖）
- Docker Compose 或設定檔內容（需讀原始碼，須學生授權）
- 報告、文件、簡報品質
- 功能邏輯正確性（需 E2E 操作測試）

# 任務
根據以下評分表原始文字，完成兩件事：
1. 萃取所有評分項目，轉為 JSON 列表。
2. 針對每一個評分項目，依據上述平台能力，判斷其「可自動偵測性」：
   - "auto"：可完全透過系統自動偵測，無需人工介入
   - "partial"：需要部分人工輔助或額外授權才能偵測
   - "manual"：完全需要人工評閱，系統無法偵測

# 輸出格式
只輸出合法的 JSON，不要有任何說明文字或 markdown。結構如下：
{
  "items": [
    {
      "id": "item-1",
      "title": "評分項目名稱",
      "description": "評分說明（從原文萃取或精簡改寫）",
      "max_score": 數字,
      "detectable": "auto | partial | manual",
      "detection_method": "若 auto/partial，具體說明偵測方式（e.g., TCP Port 80 探測）；否則 null",
      "fallback": "若 manual/partial，說明替代方案（e.g., 請學生提交截圖、要求提交 GitHub 連結）；否則 null"
    }
  ],
  "summary": "整體評分表說明，約 2-3 句，使用繁體中文。包含總配分、可自動偵測比例、哪些項目需特別注意或無法自動評分。"
}
""".strip()


async def analyze_rubric(raw_text: str) -> tuple[RubricAnalysis, dict]:
    """Send raw document text to AI, return structured RubricAnalysis."""
    if not settings.vllm_model_name:
        raise HTTPException(status_code=503, detail="VLLM_MODEL_NAME 未設定。")

    user_content = f"# 評分表原文\n\n{raw_text}"

    payload = _apply_thinking_control({
        "model": settings.vllm_model_name,
        "messages": [
            {"role": "system", "content": _ANALYZE_SYSTEM_PROMPT},
            {"role": "user", "content": user_content},
        ],
        "max_tokens": settings.vllm_max_tokens,
        "temperature": 0.2,
        "top_p": settings.vllm_top_p,
        "response_format": {"type": "json_object"},
    })

    content, metrics = await _call_vllm(payload, timeout=float(settings.vllm_timeout))

    try:
        data = json.loads(content)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"AI 回傳 JSON 解析失敗：{exc}") from exc

    items_raw = data.get("items") or []
    items: list[RubricItem] = []
    for i, raw in enumerate(items_raw):
        item_id = str(raw.get("id") or f"item-{i + 1}")
        items.append(RubricItem(
            id=item_id,
            title=str(raw.get("title") or "未命名項目"),
            description=str(raw.get("description") or ""),
            max_score=float(raw.get("max_score") or 0),
            detectable=str(raw.get("detectable") or "manual"),
            detection_method=raw.get("detection_method"),
            fallback=raw.get("fallback"),
        ))

    total_score = sum(item.max_score for item in items)
    auto_count = sum(1 for item in items if item.detectable == "auto")
    partial_count = sum(1 for item in items if item.detectable == "partial")
    manual_count = sum(1 for item in items if item.detectable == "manual")

    analysis = RubricAnalysis(
        items=items,
        total_score=total_score,
        auto_count=auto_count,
        partial_count=partial_count,
        manual_count=manual_count,
        summary=str(data.get("summary") or ""),
        raw_text=raw_text,
    )
    return analysis, metrics


# ──────────────────────────────────────────────────────────────
# 2. Chat to refine rubric
# ──────────────────────────────────────────────────────────────

_CHAT_SYSTEM_TEMPLATE = """
# 角色
你是一位專業的教學評分助理，服務對象是校園雲端平台的授課老師。
老師已上傳了一份評分表，你已完成初步分析。

# 平台背景（內部知識）
學生作業在 Proxmox VM / LXC 環境中執行（本地校園雲端，非公有雲）。
系統可偵測：Port 監聽、服務狀態、CPU/記憶體/磁碟、檔案存在、HTTP 狀態碼。
系統無法偵測：程式碼品質、DB 內容、設定檔、截圖、報告品質。

# 可用資訊來源
- 評分表結構（見下方 JSON）
- 未來將支援讀取學生 README、程式碼片段以輔助判斷（功能開發中）

# 目前評分表（JSON 格式）
{rubric_context}

# 任務
老師可能會要求修改某個評分項目的說明或配分、調整可偵測性判斷、新增或刪除項目、或詢問特定評分點的建議偵測方式。
若老師詢問的項目涉及無法自動偵測的內容，請主動說明并給出替代方案。

# 輸出格式
必須輸出合法 JSON，不要任何 markdown 包裟或自然語言。結構:
{{
  "reply": "你的回復文字（繁體中文，精簡說明修改內容或建議）",
  "updated_items": null
}}

IMPORTANT RULES:
- 如果只是回答問題或給出建議、未變更評分表結構：updated_items 設為 null。
- 如果有任何項目被新增、修改或刪除：updated_items 必須是「修改後的完整評分項目列表」（包含未修改的項目）。
- 每個 item 需有: id, title, description, max_score, detectable, detection_method, fallback。
- 不要在 reply 中重新列出所有項目，只說明變動的部分。
""".strip()


async def chat_with_rubric(
    messages: list[ChatMessage],
    rubric_context: str,
) -> tuple[str, list | None, dict]:
    """
    Multi-turn chat with rubric context injected into system prompt.
    Returns (reply_text, updated_items_or_None, metrics).
    - updated_items: complete list of RubricItem dicts when AI modified the rubric;
      None when AI only answered a question without changes.
    """
    if not settings.vllm_model_name:
        raise HTTPException(status_code=503, detail="VLLM_MODEL_NAME 未設定。")

    system_prompt = _CHAT_SYSTEM_TEMPLATE.replace(
        "{{rubric_context}}", rubric_context or "（尚未上傳評分表）"
    )

    formatted = [{"role": "system", "content": system_prompt}]
    for msg in messages:
        formatted.append({"role": msg.role, "content": msg.content})

    payload = _apply_thinking_control({
        "model": settings.vllm_model_name,
        "messages": formatted,
        "max_tokens": settings.vllm_chat_max_tokens,
        "temperature": settings.vllm_chat_temperature,
        "top_p": settings.vllm_top_p,
        "top_k": settings.vllm_top_k,
        "repetition_penalty": settings.vllm_repetition_penalty,
        "response_format": {"type": "json_object"},
    })

    content, metrics = await _call_vllm(payload, timeout=float(settings.vllm_timeout))

    # 解析結構化 JSON 回復
    reply_text = content  # fallback
    updated_items: list | None = None
    try:
        parsed = json.loads(content)
        reply_text = str(parsed.get("reply") or content)
        raw_updated = parsed.get("updated_items")
        if isinstance(raw_updated, list) and len(raw_updated) > 0:
            updated_items = raw_updated
    except (json.JSONDecodeError, TypeError):
        pass  # AI 未輸出合法 JSON，直接用原始內容作為回復

    return reply_text, updated_items, metrics


# ──────────────────────────────────────────────────────────────
# 3. Export to Excel
# ──────────────────────────────────────────────────────────────

_DETECTABLE_LABELS = {
    "auto": "✅ 可自動偵測",
    "partial": "⚠️ 部分可偵測",
    "manual": "❌ 需人工評閱",
}

_DETECTABLE_COLORS = {
    "auto": "D8F5E1",    # 綠
    "partial": "FFF3CD", # 黃
    "manual": "FDDEDE",  # 紅
}


def export_to_excel(items: list[RubricItem], summary: str = "") -> bytes:
    """Generate an .xlsx file from RubricItem list, return as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "評分表"

    # ── Header ──
    header_font = Font(bold=True, size=11)
    headers = [
        "項目編號", "評分項目", "說明",
        "配分", "可偵測性", "自動偵測方式", "替代建議",
    ]
    col_widths = [10, 25, 40, 8, 18, 35, 35]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="D0D0D0")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # ── Data rows ──
    for row_idx, item in enumerate(items, start=2):
        detectable = item.detectable or "manual"
        label = _DETECTABLE_LABELS.get(detectable, detectable)
        bg_color = _DETECTABLE_COLORS.get(detectable, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg_color)

        values = [
            item.id,
            item.title,
            item.description,
            item.max_score,
            label,
            item.detection_method or "",
            item.fallback or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 40

    # ── Summary row ──
    if summary:
        last_row = len(items) + 3
        ws.cell(row=last_row, column=1, value="備註").font = Font(bold=True)
        summary_cell = ws.cell(row=last_row, column=2, value=summary)
        ws.merge_cells(
            start_row=last_row, start_column=2,
            end_row=last_row, end_column=len(headers),
        )
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[last_row].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
